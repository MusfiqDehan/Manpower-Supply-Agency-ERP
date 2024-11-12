import calendar
import openpyxl
from openpyxl.utils import get_column_letter
from datetime import datetime
from django.shortcuts import render
from django.utils.timezone import now
from django.db.models import Sum
from django.template.loader import render_to_string
from django.http import JsonResponse, HttpResponse
from django.contrib.auth.decorators import login_required

from company_cost_app.forms import CompanyCostForm
from company_cost_app.models import CompanyCost, CostCategory


@login_required
def company_cost(request):
    user_group = request.user.Group_Users.all().first()

    if request.user.is_superuser:
        company_cost_list = CompanyCost.objects.all().order_by("-id")
        cost_categories = CostCategory.objects.all()

        # Calculate total amount of current month
        current_month = now().month
        current_month_name = calendar.month_name[now().month]
        current_year = now().year
        current_month_total = (
            CompanyCost.objects.filter(
                date_incurred__month=current_month, date_incurred__year=current_year
            ).aggregate(Sum("amount"))["amount__sum"]
            or 0
        )

        # Calculate total amount
        total_amount = CompanyCost.objects.aggregate(Sum("amount"))["amount__sum"] or 0

        filtered_amount = total_amount

        context = {
            "page_name": "Company Cost",
            "company_cost_nav_link_status": "active",
            "company_cost_list": company_cost_list,
            "cost_categories": cost_categories,
            "current_month_name": current_month_name,
            "current_month_total": current_month_total,
            "total_amount": total_amount,
            "filtered_amount": filtered_amount,
        }

        return render(request, "company_cost.html", context)
    else:
        if user_group is None:
            return render(request, "errors/403.html")
        else:
            group_permissions = user_group.permissions.all()
            for perm in group_permissions:
                if perm.codename == "cost_view":
                    company_cost_list = CompanyCost.objects.all().order_by("-id")
                    cost_categories = CostCategory.objects.all()

                    # Calculate total amount of current month
                    current_month = now().month
                    current_month_name = calendar.month_name[now().month]
                    current_year = now().year
                    current_month_total = (
                        CompanyCost.objects.filter(
                            date_incurred__month=current_month,
                            date_incurred__year=current_year,
                        ).aggregate(Sum("amount"))["amount__sum"]
                        or 0
                    )

                    # Calculate total amount
                    total_amount = (
                        CompanyCost.objects.aggregate(Sum("amount"))["amount__sum"] or 0
                    )

                    filtered_amount = total_amount

                    context = {
                        "page_name": "Company Cost",
                        "company_cost_nav_link_status": "active",
                        "company_cost_list": company_cost_list,
                        "cost_categories": cost_categories,
                        "current_month_name": current_month_name,
                        "current_month_total": current_month_total,
                        "total_amount": total_amount,
                        "filtered_amount": filtered_amount,
                    }

                    return render(request, "company_cost.html", context)
    return render(request, "errors/403.html")


@login_required
def filter_by_date(request):
    start_date = request.GET.get("start_date")
    end_date = request.GET.get("end_date")

    if start_date and end_date:
        start_date = datetime.strptime(start_date, "%Y-%m-%d")
        end_date = datetime.strptime(end_date, "%Y-%m-%d")
        company_cost_list = CompanyCost.objects.filter(
            date_incurred__range=(start_date, end_date)
        ).order_by("-id")
        filtered_amount = (
            CompanyCost.objects.filter(
                date_incurred__range=(start_date, end_date)
            ).aggregate(Sum("amount"))["amount__sum"]
            or 0
        )
    else:
        company_cost_list = CompanyCost.objects.all().order_by("-id")
        filtered_amount = (
            CompanyCost.objects.aggregate(Sum("amount"))["amount__sum"] or 0
        )

    data = {
        "filtered_amount": filtered_amount,
        "company_cost_list": render_to_string(
            "company_cost_list.html",
            {"company_cost_list": company_cost_list},
            request=request,
        ),
    }

    return JsonResponse(data)


@login_required
def export_company_cost(request):
    start_date = request.GET.get("start_date")
    end_date = request.GET.get("end_date")

    if start_date and end_date:
        start_date = datetime.strptime(start_date, "%Y-%m-%d")
        end_date = datetime.strptime(end_date, "%Y-%m-%d")
        company_cost_list = CompanyCost.objects.filter(
            date_incurred__range=(start_date, end_date)
        ).order_by("-id")
    else:
        company_cost_list = CompanyCost.objects.all().order_by("-id")

    # Create an in-memory output file for the new workbook.
    output = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    output["Content-Disposition"] = "attachment; filename=company_costs.xlsx"

    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = "Company Costs"

    # Define the titles for columns
    columns = ["Date Incurred", "Note", "Amount"]
    row_num = 1

    # Assign the titles for each cell of the header
    for col_num, column_title in enumerate(columns, 1):
        cell = worksheet.cell(row=row_num, column=col_num)
        cell.value = column_title

    # Iterate through all company costs and write them to the sheet
    for company_cost in company_cost_list:
        row_num += 1
        row = [
            company_cost.date_incurred,
            company_cost.note,
            company_cost.amount,
        ]
        for col_num, cell_value in enumerate(row, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = cell_value

    workbook.save(output)
    return output


# View to handle AJAX requests for category filtering
def filter_by_category(request):
    category_id = request.GET.get("category")

    if category_id:
        company_cost_list = CompanyCost.objects.filter(
            category_id=category_id
        ).order_by("-id")
        # Handle filtering logic based on the selected category
        filtered_amount = (
            CompanyCost.objects.filter(category_id=category_id).aggregate(
                Sum("amount")
            )["amount__sum"]
            or 0
        )
    else:
        company_cost_list = CompanyCost.objects.all().order_by("-id")
        total_amount = CompanyCost.objects.aggregate(Sum("amount"))["amount__sum"] or 0
        filtered_amount = total_amount

    data = {
        "filtered_amount": filtered_amount,
        "company_cost_list": render_to_string(
            "company_cost_list.html",
            {"company_cost_list": company_cost_list},
            request=request,
        ),
    }

    return JsonResponse(data)


@login_required
def company_cost_create(request):
    data = dict()
    user_group = request.user.Group_Users.all().first()

    if request.user.is_superuser:
        cost_categories = CostCategory.objects.all()

        if request.method == "POST":
            form = CompanyCostForm(request.POST, request.FILES)

            if form.is_valid():
                form.save()

                # Calculate updated filtered amount
                category_id = request.POST.get("category")
                if category_id:
                    filtered_amount = (
                        CompanyCost.objects.filter(category_id=category_id).aggregate(
                            Sum("amount")
                        )["amount__sum"]
                        or 0
                    )
                else:
                    filtered_amount = (
                        CompanyCost.objects.aggregate(Sum("amount"))["amount__sum"] or 0
                    )

                # Calculate updated totals
                current_month = now().month
                current_year = now().year
                current_month_total = (
                    CompanyCost.objects.filter(
                        date_incurred__month=current_month,
                        date_incurred__year=current_year,
                    ).aggregate(Sum("amount"))["amount__sum"]
                    or 0
                )

                total_amount = (
                    CompanyCost.objects.aggregate(Sum("amount"))["amount__sum"] or 0
                )

                data["form_is_valid"] = True
                company_cost_list = CompanyCost.objects.all().order_by("-id")

                context = {
                    "company_cost_list": company_cost_list,
                    "user": request.user,
                }
                data["company_cost_list"] = render_to_string(
                    "company_cost_list.html", context, request=request
                )
                data["current_month_total"] = current_month_total
                data["filtered_amount"] = filtered_amount
                data["total_amount"] = total_amount
                return JsonResponse(data)

            else:
                data["form_is_valid"] = False
                print(form.errors)
                return JsonResponse(data)

        else:
            form = CompanyCostForm()  # Get Request

        context = {
            "form": form,
            "user": request.user,
            "cost_categories": cost_categories,
        }

        data["html_form"] = render_to_string(
            "company_cost_create.html", context, request=request
        )

        return JsonResponse(data)
    else:
        if user_group is None:
            return render(request, "errors/403-contents.html")
        else:
            group_permissions = user_group.permissions.all()
            for perm in group_permissions:
                if perm.codename == "cost_create":
                    cost_categories = CostCategory.objects.all()

                    if request.method == "POST":
                        form = CompanyCostForm(request.POST, request.FILES)

                        if form.is_valid():
                            form.save()

                            # Calculate updated filtered amount
                            category_id = request.POST.get("category")
                            if category_id:
                                filtered_amount = (
                                    CompanyCost.objects.filter(
                                        category_id=category_id
                                    ).aggregate(Sum("amount"))["amount__sum"]
                                    or 0
                                )
                            else:
                                filtered_amount = (
                                    CompanyCost.objects.aggregate(Sum("amount"))[
                                        "amount__sum"
                                    ]
                                    or 0
                                )

                            # Calculate updated totals
                            current_month = now().month
                            current_year = now().year
                            current_month_total = (
                                CompanyCost.objects.filter(
                                    date_incurred__month=current_month,
                                    date_incurred__year=current_year,
                                ).aggregate(Sum("amount"))["amount__sum"]
                                or 0
                            )

                            total_amount = (
                                CompanyCost.objects.aggregate(Sum("amount"))[
                                    "amount__sum"
                                ]
                                or 0
                            )

                            data["form_is_valid"] = True
                            company_cost_list = CompanyCost.objects.all().order_by(
                                "-id"
                            )

                            context = {
                                "company_cost_list": company_cost_list,
                                "user": request.user,
                            }
                            data["company_cost_list"] = render_to_string(
                                "company_cost_list.html", context, request=request
                            )
                            data["current_month_total"] = current_month_total
                            data["filtered_amount"] = filtered_amount
                            data["total_amount"] = total_amount
                            return JsonResponse(data)

                        else:
                            data["form_is_valid"] = False
                            print(form.errors)
                            return JsonResponse(data)

                    else:
                        form = CompanyCostForm()  # Get Request

                    context = {
                        "form": form,
                        "user": request.user,
                        "cost_categories": cost_categories,
                    }

                    data["html_form"] = render_to_string(
                        "company_cost_create.html", context, request=request
                    )

                    return JsonResponse(data)
    data["html_form"] = render_to_string(
        "errors/403-contents.html", {}, request=request
    )
    return JsonResponse(data)


# ============Company Cost Edit Page.===========
@login_required
def company_cost_edit(request, pk):
    data = dict()
    company_cost = CompanyCost.objects.get(pk=pk)
    user_group = request.user.Group_Users.all().first()

    if request.user.is_superuser:
        if request.method == "POST":
            form = CompanyCostForm(request.POST, request.FILES, instance=company_cost)
            if form.is_valid():
                form.save()
                data["form_is_valid"] = True

                # Calculate updated filtered amount
                category_id = request.POST.get("category")
                if category_id:
                    filtered_amount = (
                        CompanyCost.objects.filter(category_id=category_id).aggregate(
                            Sum("amount")
                        )["amount__sum"]
                        or 0
                    )
                else:
                    filtered_amount = (
                        CompanyCost.objects.aggregate(Sum("amount"))["amount__sum"] or 0
                    )

                # Calculate updated totals
                current_month = now().month
                current_year = now().year
                current_month_total = (
                    CompanyCost.objects.filter(
                        date_incurred__month=current_month,
                        date_incurred__year=current_year,
                    ).aggregate(Sum("amount"))["amount__sum"]
                    or 0
                )
                total_amount = (
                    CompanyCost.objects.aggregate(Sum("amount"))["amount__sum"] or 0
                )

                company_cost_list = CompanyCost.objects.all().order_by("-id")
                context = dict(company_cost_list=company_cost_list)

                data["company_cost_list"] = render_to_string(
                    "company_cost_list.html", context, request=request
                )

                data["current_month_total"] = current_month_total
                data["filtered_amount"] = filtered_amount
                data["total_amount"] = total_amount

                return JsonResponse(data)
            else:
                data["form_is_valid"] = False
                return JsonResponse(data)

        else:
            print("Cost Edit Get Request")
            # cost_categories = CostCategory.objects.all()

            if company_cost.cost_document:
                context = dict(company_cost=company_cost)

                data["html_form"] = render_to_string(
                    "company_cost_edit.html", context, request=request
                )

            else:
                company_cost.cost_document = "default.pdf"
                company_cost.save()
                context = dict(company_cost=company_cost)

                data["html_form"] = render_to_string(
                    "company_cost_edit.html", context, request=request
                )

            return JsonResponse(data)
    else:
        if user_group is None:
            return render(request, "errors/403-contents.html")
        else:
            group_permissions = user_group.permissions.all()
            for perm in group_permissions:
                if perm.codename == "cost_update":
                    if request.method == "POST":
                        form = CompanyCostForm(
                            request.POST, request.FILES, instance=company_cost
                        )
                        if form.is_valid():
                            form.save()
                            data["form_is_valid"] = True

                            # Calculate updated filtered amount
                            category_id = request.POST.get("category")
                            if category_id:
                                filtered_amount = (
                                    CompanyCost.objects.filter(
                                        category_id=category_id
                                    ).aggregate(Sum("amount"))["amount__sum"]
                                    or 0
                                )
                            else:
                                filtered_amount = (
                                    CompanyCost.objects.aggregate(Sum("amount"))[
                                        "amount__sum"
                                    ]
                                    or 0
                                )

                            # Calculate updated totals
                            current_month = now().month
                            current_year = now().year
                            current_month_total = (
                                CompanyCost.objects.filter(
                                    date_incurred__month=current_month,
                                    date_incurred__year=current_year,
                                ).aggregate(Sum("amount"))["amount__sum"]
                                or 0
                            )
                            total_amount = (
                                CompanyCost.objects.aggregate(Sum("amount"))[
                                    "amount__sum"
                                ]
                                or 0
                            )

                            company_cost_list = CompanyCost.objects.all().order_by(
                                "-id"
                            )
                            context = dict(company_cost_list=company_cost_list)

                            data["company_cost_list"] = render_to_string(
                                "company_cost_list.html", context, request=request
                            )

                            data["current_month_total"] = current_month_total
                            data["filtered_amount"] = filtered_amount
                            data["total_amount"] = total_amount

                            return JsonResponse(data)
                        else:
                            data["form_is_valid"] = False
                            return JsonResponse(data)

                    else:
                        cost_categories = CostCategory.objects.all()
                        context = dict(
                            company_cost=company_cost, cost_categories=cost_categories
                        )
                        data["html_form"] = render_to_string(
                            "company_cost_edit.html", context, request=request
                        )

                        return JsonResponse(data)
    data["html_form"] = render_to_string(
        "errors/403-contents.html", {}, request=request
    )
    return JsonResponse(data)


# ============Company Cost Delete Page.===========
@login_required
def company_cost_delete(request, pk):
    data = dict()
    company_cost = CompanyCost.objects.get(pk=pk)
    user_group = request.user.Group_Users.all().first()

    if request.user.is_superuser:
        if request.method == "POST":
            company_cost.delete()
            data["form_is_valid"] = True

            # Calculate updated filtered amount
            category_id = request.POST.get("category")
            if category_id:
                filtered_amount = (
                    CompanyCost.objects.filter(category_id=category_id).aggregate(
                        Sum("amount")
                    )["amount__sum"]
                    or 0
                )
            else:
                filtered_amount = (
                    CompanyCost.objects.aggregate(Sum("amount"))["amount__sum"] or 0
                )

            # Calculate updated totals
            current_month = now().month
            current_year = now().year
            current_month_total = (
                CompanyCost.objects.filter(
                    date_incurred__month=current_month, date_incurred__year=current_year
                ).aggregate(Sum("amount"))["amount__sum"]
                or 0
            )
            total_amount = (
                CompanyCost.objects.aggregate(Sum("amount"))["amount__sum"] or 0
            )

            company_cost_list = CompanyCost.objects.all().order_by("-id")
            context = dict(company_cost_list=company_cost_list)
            data["company_cost_list"] = render_to_string(
                "company_cost_list.html", context, request=request
            )

            data["current_month_total"] = current_month_total
            data["filtered_amount"] = filtered_amount
            data["total_amount"] = total_amount

            return JsonResponse(data)
        else:
            context = dict(company_cost=company_cost)
            data["html_form"] = render_to_string(
                "company_cost_delete.html", context, request=request
            )

            return JsonResponse(data)
    else:
        if user_group is None:
            return render(request, "errors/403-contents.html")
        else:
            group_permissions = user_group.permissions.all()
            for perm in group_permissions:
                if perm.codename == "cost_delete":
                    if request.method == "POST":
                        company_cost.delete()
                        data["form_is_valid"] = True

                        # Calculate updated filtered amount
                        category_id = request.POST.get("category")
                        if category_id:
                            filtered_amount = (
                                CompanyCost.objects.filter(
                                    category_id=category_id
                                ).aggregate(Sum("amount"))["amount__sum"]
                                or 0
                            )
                        else:
                            filtered_amount = (
                                CompanyCost.objects.aggregate(Sum("amount"))[
                                    "amount__sum"
                                ]
                                or 0
                            )

                        # Calculate updated totals
                        current_month = now().month
                        current_year = now().year
                        current_month_total = (
                            CompanyCost.objects.filter(
                                date_incurred__month=current_month,
                                date_incurred__year=current_year,
                            ).aggregate(Sum("amount"))["amount__sum"]
                            or 0
                        )
                        total_amount = (
                            CompanyCost.objects.aggregate(Sum("amount"))["amount__sum"]
                            or 0
                        )

                        company_cost_list = CompanyCost.objects.all().order_by("-id")
                        context = dict(company_cost_list=company_cost_list)
                        data["company_cost_list"] = render_to_string(
                            "company_cost_list.html", context, request=request
                        )

                        data["current_month_total"] = current_month_total
                        data["filtered_amount"] = filtered_amount
                        data["total_amount"] = total_amount

                        return JsonResponse(data)
                    else:
                        context = dict(company_cost=company_cost)
                        data["html_form"] = render_to_string(
                            "company_cost_delete.html", context, request=request
                        )

                        return JsonResponse(data)
    data["html_form"] = render_to_string(
        "errors/403-contents.html", {}, request=request
    )
    return JsonResponse(data)
